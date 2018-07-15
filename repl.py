import xlrd
#import xlwt
import os
# import re
import re
import wx
#import time
from xlutils.copy import copy
from openpyxl import load_workbook

global target_file
target_file = r''

class MainWindows(wx.Frame):
	def __init__(self,parent,title):
		wx.Frame.__init__(self,None,-1,title=title,size=(400,200),style=wx.CAPTION|wx.CLOSE_BOX)
		# self.icon=wx.Icon('W.ico', type=wx.BITMAP_TYPE_ICO)
		# self.SetIcon(self.icon)
		panel = wx.Panel(self)
		self.label = wx.StaticText(panel,-1,'选择文件：',size=(60,20),pos = (10,0))
		self.text = wx.TextCtrl(panel,-1,target_file,size=(200,20),pos = (70,0))
		self.open_dir_button = wx.Button(panel,-1,'选择文件',size = (90,20),pos=(280,0))
		self.run_button = wx.Button(panel,-1,'测试按钮',size=(80,20),pos = (20,40))
		self.Status_Bar = self.CreateStatusBar(number =1)
	

		# target_file = r'E:\百度云同步盘\LD\桌面备份'
		self.run_button.Bind(wx.EVT_BUTTON,self.OnRunButtonClick)
		self.open_dir_button.Bind(wx.EVT_BUTTON,self.OnOpenDirButtonClick)

	# def OnPaintMotion(self,event):
	# 	self.statusbar.SetStatusText(u"鼠标位置：" + str(event.GetPositionTuple()), 0)
	# 	self.Skip()

	def Set_Status_Bar(self,T):
		self.Status_Bar.SetStatusText(T)

	def OnOpenDirButtonClick(self,event):
		self.Status_Bar.SetStatusText('Open a xlsx files')
		dlg = wx.FileDialog(self,u"选择文件",wildcard = "XLSX files(*.xlsx)|*.xlsx",style=wx.FD_OPEN)
		if dlg.ShowModal()==wx.ID_OK:
			target_file = dlg.GetPath()
			self.text.SetValue(target_file)
			print(target_file)
		dlg.Destroy()

	def repl(self,target_file):

		data = xlrd.open_workbook(target_file,formatting_info=False)
		table = data.sheet_by_name(u'数据库')
		mod_book = copy(data)
		mod_sheet = mod_book.get_sheet(u'数据库')
		wb = load_workbook(target_file)
		# sheet_ranges = wb['数据库']
		ws =wb['数据库']


		# print (table.cell(0,10).value)
		# print (table.cell(0,1).value)
		for k in range(table.ncols):
			if table.cell(0,k).value == u'货物代码':
				hwdm_col = k
				print(hwdm_col)
			elif table.cell(0,k).value == u'进出口要素':
				jckys_col = k
				print(jckys_col)
		# print(hwdm_col,jckys_col)
		# # print(table.nrows)
		#第一遍检索
		h = 0
		for i in range(table.nrows):
			hwdm = str(table.cell(i,hwdm_col).value).split('.')[0]
			jckys = str(table.cell(i,jckys_col).value)
			pattern = r'\d{5,12}'
			# print(hwdm)
			if not re.findall(u'进出口要素',jckys):
				if not re.findall(u'无GTIN',jckys):
			 		if not re.findall(u'无CAS',jckys):
			 			if not re.findall(u'无其他申报要素',jckys):
			 				jckys = jckys+'|无GTIN|无CAS|无其他申报要素'
			 				ws.cell(row=i+1,column=jckys_col+2).value = jckys
			 				print(jckys)
			s = re.search(pattern,jckys)
			if not s==None:
				if s.group(0) == hwdm:
			 		continue
				# num = s.span()
				# print(jckys[:num[0]])
				# print(jckys[num[1]:])
			ex = re.findall(pattern,jckys)
			for x in ex:
				jckys=jckys.replace(x,hwdm)
			# print(jckys)

				# ex = ' '.join(ex)
				# print('ex = ' +ex)
				# jckys.replace()		
				#newtext = jckys[:num[0]] + hwdm + jckys[num[1]:]
					
				#print(newtext)
				# mod_sheet.write(i,jckys_col+1,newtext)
				ws.cell(row=i+1,column=jckys_col+2).value = jckys
				
				# print(jckys.group())
				# print(jckys.span())

				# print(jckys)
				# repl(hwdm,jckys)


				h+=1
				# print(h)
				print('已处理第'+str(i+1)+'行'+'     '+jckys)
				self.Set_Status_Bar(str(i+1))
			#time.sleep(1)
			# frame.Status_Bar.SetStatusText(str(i+1))
			# Status_Bar.SetStatusText(u"鼠标位置：" + str(event.GetPositionTuple()), 0)
		# mod_book.save(r'E:\desktop\测试\1.xls')
		wb.save(target_file)

		dlg=wx.MessageDialog(None,u"处理完成！","完成！",style=wx.OK|wx.ICON_INFORMATION)
		result=dlg.ShowModal()
		dlg.Destroy()

	def OnRunButtonClick(self,event):	
		target_file = self.text.GetValue()
		if not os.path.exists(target_file):
			dlg=wx.MessageDialog(self,u"没有选择文件！","错误！",style=wx.OK|wx.ICON_ERROR)
			result=dlg.ShowModal()
			dlg.Destroy()
		else:
			self.repl(target_file)


def main():
	app = wx.App(False)
	frame = MainWindows(None,title = "替换商品号")
	frame.Show(True)
	app.MainLoop()

if __name__ == '__main__':
	main()